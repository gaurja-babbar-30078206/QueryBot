import pandas as pd
from functools import reduce
import openpyxl
import evaluate
import numpy as np
import pandas as pd
from langchain_huggingface import HuggingFaceEmbeddings
from sklearn.metrics.pairwise import cosine_similarity
from rouge_score import rouge_scorer



"""
Task
- to make a new excel sheet
- Adding nine columns for the answers
- after that nine columns for the cosine similarity
- for cosine similarity
- reference embeddings
- candidate embeddings

Now to add Recall score for each question and answer

 
"""

removeWords = ["\nAssistant:", "\nBot:", "\nHuman:", '\nSystem']


def get_formatted_answers(answers_per_sheet:list):
    final_answer_list = []
    for answers in answers_per_sheet:
           answers = str(answers)
           for sub in removeWords:
               answers = answers.replace(sub, "")
           final_answer_list.append(answers)
    return final_answer_list
    
    
def calc_rouge_score_eval_library(path, reference):
    df = pd.read_excel(path, sheet_name= None)
    sheets = list(df.keys())
    results = {}
    for sheet_name in sheets:
        anwsers =  df[sheet_name]["Answers"]
        predictions = get_formatted_answers(anwsers)
        rouge = evaluate.load('rouge')
        results[sheet_name] = rouge.compute(predictions= predictions, references= reference)
    print(results)   
       
    
          
          
       
        
def get_reference_list():
    path = r"D:\OneDrive - Adani\Desktop\LEARNING_FOLDER\_Kolkata_2024\1_LLM\3_Text_query_bot\_docs\benchmark_qa_3_temp_zero\1+gpt35turbo+snowflake-arctic-embed-m.xlsx"

    wb_obj = openpyxl.load_workbook(path)
    sheets = wb_obj.sheetnames
    df = pd.read_excel(path, sheet_name= sheets)
    keys = df.keys()

    final_answer_list = []
    first_list = df["snowflakeArctic"]["Answers"]
    second_list = df["bgebase"]["Answers"]
    third_list = df["gteBase"]["Answers"]


    for i in range(20):
        # print(i)
        test = []
        test.append(first_list[i])    
        test.append(second_list[i])    
        test.append(third_list[i])
        final_answer_list.append(test)

    references = final_answer_list
    return references       
    
    
def get_ref_embeddins():
    path = r"D:\OneDrive - Adani\Desktop\LEARNING_FOLDER\_Kolkata_2024\1_LLM\3_Text_query_bot\streamlit_examples\gpt_qa_embeddings.npy"
    return np.load(path)

def init_embeddings():
    dir= r"D:\OneDrive - Adani\Desktop\LEARNING_FOLDER\_Kolkata_2024\1_LLM\local_downloaded_models\embedding_models\snowflake-arctic-embed-m"
    return  HuggingFaceEmbeddings(
            model_name = dir,
            show_progress = True,
            model_kwargs = {"trust_remote_code": True})


def calc_cosine_sim(ref, cand):
    # print(f"Embedding --->{cand}")
    # print(f" REf Embedding --->{ref}")
    cosine_score_list = []
    for i in range(20):
        score = cosine_similarity([ref[i]], [cand[i]])
        cosine_score_list.append(score[0][0])
    return cosine_score_list    


def edit_sheet(sheet, score, cell:str):
    for i in range(2,22):
        sheet[f"{cell}{i}"] = score[i-2]    
        
def calc_cosine_sim(ref):
    path = r"D:\OneDrive - Adani\Desktop\LEARNING_FOLDER\_Kolkata_2024\1_LLM\3_Text_query_bot\_docs\benchmark_qa_3_temp_zero\llm_responses_document_query_bot.xlsx"
    xls = pd.ExcelFile(path)
    sheets = xls.sheet_names
    count = 0
    # embedding_set = set()
    test_wb = openpyxl.load_workbook(path)
    for sheet in sheets:
        df = pd.read_excel(path, sheet_name=sheet)
        answers = get_formatted_answers(df["Answers"])
        embed_llm = init_embeddings()
        answer_embedding = embed_llm.embed_documents(answers)
        print(f"Answer embedding ---> {len(answer_embedding)}")
        print(f"Length of Answers -----> {len(answers)}")
        count += 1
        score = calc_cosine_sim( ref= ref, cand= answer_embedding)
        print(f"Score of sheet ----> {count} {score}")
        edit_sheet(test_wb[sheet],score)
    
    test_wb.save(path)    
        
    print(f"This is count ---------------> {count}")


metric_list = ['rouge1_recall','rouge1_precision','rouge1_fmeasure',
               'rouge2_recall','rouge2_precision','rouge2_fmeasure',
               'rougeL_recall','rougeL_precision','rougeL_fmeasure',
               'rougeLsum_recall','rougeLsum_precision','rougeLsum_fmeasure']


def make_dynamic_list():
    final_list = []
    for i in range(len(metric_list)):
        final_list.append([])
    return final_list


    
def recall_prec_fmeasure():
    
    reference = get_reference_list()
    path = r"D:\OneDrive - Adani\Desktop\LEARNING_FOLDER\_Kolkata_2024\1_LLM\3_Text_query_bot\_docs\benchmark_qa_3_temp_zero\llm_responses_document_query_bot.xlsx"
    xls = pd.ExcelFile(path)
    sheets = xls.sheet_names
    count = 0
    
    test_wb = openpyxl.load_workbook(path)
    
    for sheet in sheets:
        all_rouge_metric_list = make_dynamic_list()
        
        df = pd.read_excel(path, sheet_name=sheet)
        answers = get_formatted_answers(df["Answers"])
        # print(f"Answers -----> {answers}")
        print(f"Answers Length-----> {len(answers)} -------------Sheet {sheet}")
        for i in range(20):
            # score = rouge_scorer(reference[i], answers[i])
            scorer = rouge_scorer.RougeScorer(rouge_types= [ 'rouge1','rouge2', 'rougeL', 'rougeLsum'])
            score = scorer.score(target= str(reference[i][0]), prediction= str(answers[i]))
            # print(f"REf ---->{str(reference[i])}")
            # print(f"Candidate ---->{str(answers[i])}")
            print(f"Rouge Score ------ {score}----- SHEET ---- {sheet}")
            all_rouge_metric_list[0].append(score['rouge1'].recall)
            all_rouge_metric_list[1].append(score['rouge1'].precision)
            all_rouge_metric_list[2].append(score['rouge1'].fmeasure)
            all_rouge_metric_list[3].append(score['rouge2'].recall)
            all_rouge_metric_list[4].append(score['rouge2'].precision)
            all_rouge_metric_list[5].append(score['rouge2'].fmeasure)
            all_rouge_metric_list[6].append(score['rougeL'].recall)
            all_rouge_metric_list[7].append(score['rougeL'].precision)
            all_rouge_metric_list[8].append(score['rougeL'].fmeasure)
            all_rouge_metric_list[9].append(score['rougeLsum'].recall)
            all_rouge_metric_list[10].append(score['rougeLsum'].precision)
            all_rouge_metric_list[11].append(score['rougeLsum'].fmeasure)
        
        edit_sheet(test_wb[sheet],score= all_rouge_metric_list[0],cell='E')    
        edit_sheet(test_wb[sheet],score= all_rouge_metric_list[1],cell='F')    
        edit_sheet(test_wb[sheet],score= all_rouge_metric_list[2],cell='G')    
        edit_sheet(test_wb[sheet],score= all_rouge_metric_list[3],cell='H')    
        edit_sheet(test_wb[sheet],score= all_rouge_metric_list[4],cell='I')    
        edit_sheet(test_wb[sheet],score= all_rouge_metric_list[5],cell='J')    
        edit_sheet(test_wb[sheet],score= all_rouge_metric_list[6],cell='K')    
        edit_sheet(test_wb[sheet],score= all_rouge_metric_list[7],cell='L')    
        edit_sheet(test_wb[sheet],score= all_rouge_metric_list[8],cell='M')    
        edit_sheet(test_wb[sheet],score= all_rouge_metric_list[9],cell='N')    
        edit_sheet(test_wb[sheet],score= all_rouge_metric_list[10],cell='O')    
        edit_sheet(test_wb[sheet],score= all_rouge_metric_list[11],cell='P')    
            
        print(f"All Metric List -----> {len(all_rouge_metric_list[0])}")
        
    test_wb.save(path)    
        
    print(f"This is count ---------------> {count}")
    # get answers list 
    # iterate over 20 range
    # calc rouge score for each sentence,
    # extract recall, precision, fmeasure
    # add to the sheet


# def calc_bleu_score():
                 
if __name__ == "__main__":
    # reference  = get_reference_list()
    
    # prediction_path = r"D:\OneDrive - Adani\Desktop\LEARNING_FOLDER\_Kolkata_2024\1_LLM\3_Text_query_bot\_docs\benchmark_qa_3_temp_zero\all_llama-2-7b-chat.Q6_K.gguf_responses.xlsx"
    # prediction_path = r"D:\OneDrive - Adani\Desktop\LEARNING_FOLDER\_Kolkata_2024\1_LLM\3_Text_query_bot\_docs\benchmark_qa_2\1+Mistral-7B-Instruct-v0.3.Q4_K_M.gguf+snowflake-arctic-embed-m.xlsx"
    # prediction_path = r"D:\OneDrive - Adani\Desktop\LEARNING_FOLDER\_Kolkata_2024\1_LLM\3_Text_query_bot\_docs\benchmark_qa_2\2+Mistral-7B-Instruct-v0.3.Q4_K_M.gguf+bge-base-en-v1.5.xlsx"
    # prediction_path = r"D:\OneDrive - Adani\Desktop\LEARNING_FOLDER\_Kolkata_2024\1_LLM\3_Text_query_bot\_docs\benchmark_qa_2\3+Mistral-7B-Instruct-v0.3.Q4_K_M.gguf+gte-base-en-v1.5.xlsx"
    # prediction_path = r"D:\OneDrive - Adani\Desktop\LEARNING_FOLDER\_Kolkata_2024\1_LLM\3_Text_query_bot\_docs\benchmark_qa_2\4+llama-2-7b-chat.Q4_K_M.gguf+snowflake-arctic-embed-m.xlsx"
    # prediction_path = r"D:\OneDrive - Adani\Desktop\LEARNING_FOLDER\_Kolkata_2024\1_LLM\3_Text_query_bot\_docs\benchmark_qa_2\5+llama-2-7b-chat.Q4_K_M.gguf+bge-base-en-v1.5.xlsx"
    # prediction_path = r"D:\OneDrive - Adani\Desktop\LEARNING_FOLDER\_Kolkata_2024\1_LLM\3_Text_query_bot\_docs\benchmark_qa_2\6+llama-2-7b-chat.Q4_K_M.gguf+gte-base-en-v1.5.xlsx"
    # calc_rouge_score(prediction_path,reference)
    # can_embeddings = []    
    
    # ref_embedding = get_ref_embeddins()
    # print(len(ref_embedding[0]))
    # calc_cosine_sim(ref=ref_embedding
    # print((scores))
    # reference =get_reference_list()
    # print(len(reference))
    # print(len(reference))
    recall_prec_fmeasure()
    # test_list = make_dynamic_list()
    # print(test_list)
    

    